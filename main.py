from fastapi import FastAPI, File, UploadFile, Header, HTTPException, Body
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from io import BytesIO
import os
import re
import unicodedata
import json
import base64
import zipfile
from typing import Any

import psycopg


app = FastAPI()

API_TOKEN = os.getenv("IMPORT_API_TOKEN", "")
DATABASE_URL = os.getenv("DATABASE_URL", "")


def check_token(x_api_token: str | None):
    if not API_TOKEN:
        raise HTTPException(status_code=500, detail="Server token not configured")
    if x_api_token != API_TOKEN:
        raise HTTPException(status_code=401, detail="Unauthorized")


def get_conn():
    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="DATABASE_URL not configured")
    return psycopg.connect(DATABASE_URL, sslmode="require")


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def sql_identifier(name: str) -> str:
    clean = normalize_key(name)
    if not clean:
        clean = "col"
    if clean[0].isdigit():
        clean = f"c_{clean}"
    return clean


PRODUCT_HEADER_CANDIDATES = [
    "type_de_fournitures_services",
    "type_fournitures_services",
    "fournitures_services",
    "fourniture",
    "produit",
    "product",
    "article",
    "designation",
    "libelle",
    "description",
    "item",
    "nom_produit",
    "prestation",
    "service",
]

QTY_HEADER_CANDIDATES = [
    "quantite",
    "quantity",
    "qty",
    "qte",
    "qte_",
    "nombre",
    "nb",
    "volume",
]

LIBELLE_CANDIDATES = [
    "libelle",
    "designation",
    "designation_article",
    "designation_produit",
    "description",
    "article",
    "produit",
    "product",
    "nom_produit",
    "item",
    "service",
    "prestation",
]


def load_workbook_safe(content: bytes):
    try:
        return load_workbook(filename=BytesIO(content), data_only=True)
    except (InvalidFileException, zipfile.BadZipFile):
        raise HTTPException(
            status_code=400,
            detail="Unsupported Excel format. The uploaded file is not a valid .xlsx file."
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Unable to read Excel file: {str(e)}"
        )


def choose_best_sheet(wb):
    best_sheet_name = wb.sheetnames[0]
    best_rows = []
    best_score = -1

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        non_empty = 0
        text_cells = 0

        for row in rows[:300]:
            for cell in row:
                if cell is None or str(cell).strip() == "":
                    continue
                non_empty += 1
                if any(ch.isalpha() for ch in str(cell)):
                    text_cells += 1

        score = non_empty + text_cells * 2

        if score > best_score:
            best_score = score
            best_sheet_name = ws.title
            best_rows = rows

    return best_sheet_name, best_rows


def detect_header_row(rows: list[list[Any]]):
    best_score = -1
    best_row_idx = None
    best_mapping = {}

    for row_idx, row in enumerate(rows[:40]):
        mapping = {}
        score = 0

        for col_idx, cell in enumerate(row):
            norm = normalize_key(cell)
            if not norm:
                continue

            mapping[col_idx] = norm

            if norm in PRODUCT_HEADER_CANDIDATES:
                score += 10
            if norm in QTY_HEADER_CANDIDATES:
                score += 5
            if len(norm) > 2:
                score += 1

        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_mapping = mapping

    if best_score < 5:
        return None, {}

    return best_row_idx, best_mapping


def find_column_by_candidates(header_map: dict[int, str], candidates: list[str]):
    for candidate in candidates:
        for col_idx, norm_name in header_map.items():
            if norm_name == candidate:
                return col_idx
    return None


def fallback_detect_product_col(data_rows: list[list[Any]]):
    if not data_rows:
        return None

    max_cols = max(len(r) for r in data_rows)
    best_col = None
    best_score = -1

    for col_idx in range(max_cols):
        score = 0
        non_empty = 0
        long_text = 0

        for row in data_rows[:200]:
            val = row[col_idx] if col_idx < len(row) else None
            txt = normalize_text(val)
            if not txt:
                continue

            non_empty += 1
            if any(ch.isalpha() for ch in txt):
                score += 1
            if len(txt) >= 10:
                long_text += 1
                score += 2

        if non_empty >= 3 and long_text >= 3 and score > best_score:
            best_score = score
            best_col = col_idx

    return best_col


def find_best_libelle_column(columns: list[str]) -> str | None:
    normalized = [normalize_key(c) for c in columns]

    for candidate in LIBELLE_CANDIDATES:
        for i, col in enumerate(normalized):
            if col == candidate:
                return columns[i]

    for candidate in LIBELLE_CANDIDATES:
        for i, col in enumerate(normalized):
            if col.startswith(candidate):
                return columns[i]

    for candidate in LIBELLE_CANDIDATES:
        for i, col in enumerate(normalized):
            if candidate in col:
                return columns[i]

    return None


def parse_client_xlsx_bytes(content: bytes):
    wb = load_workbook_safe(content)
    sheet_name, rows = choose_best_sheet(wb)

    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")

    header_row_idx, header_map = detect_header_row(rows)
    parsed_rows = []

    if header_row_idx is not None:
        data_rows = rows[header_row_idx + 1 :]
        product_col = find_column_by_candidates(header_map, PRODUCT_HEADER_CANDIDATES)
        qty_col = find_column_by_candidates(header_map, QTY_HEADER_CANDIDATES)

        if product_col is None:
            product_col = fallback_detect_product_col(data_rows)

        for idx, row in enumerate(data_rows, start=1):
            product_value = row[product_col] if product_col is not None and product_col < len(row) else None
            product_label = normalize_text(product_value)

            if not product_label:
                continue

            low = product_label.lower()
            if "type de fournitures" in low:
                continue
            if "reference fournisseur" in low:
                continue
            if low == "pu":
                continue
            if len(product_label) < 3:
                continue

            quantity = 1
            if qty_col is not None and qty_col < len(row):
                raw_qty = row[qty_col]
                txt_qty = normalize_text(raw_qty).replace(",", ".")
                try:
                    q = float(txt_qty)
                    if q > 0:
                        quantity = q
                except Exception:
                    quantity = 1

            parsed_rows.append({
                "request_row_number": idx,
                "product_label": product_label,
                "product_label_clean": normalize_text(product_label).upper(),
                "quantity": quantity,
            })

        return {
            "success": True,
            "sheet_name": sheet_name,
            "header_row_index": header_row_idx + 1,
            "product_col_index": product_col + 1 if product_col is not None else None,
            "qty_col_index": qty_col + 1 if qty_col is not None else None,
            "rows_parsed": len(parsed_rows),
            "rows": parsed_rows,
        }

    data_rows = rows
    product_col = fallback_detect_product_col(data_rows)

    for idx, row in enumerate(data_rows, start=1):
        product_value = row[product_col] if product_col is not None and product_col < len(row) else None
        product_label = normalize_text(product_value)

        if not product_label:
            continue
        if len(product_label) < 3:
            continue

        parsed_rows.append({
            "request_row_number": idx,
            "product_label": product_label,
            "product_label_clean": normalize_text(product_label).upper(),
            "quantity": 1,
        })

    return {
        "success": True,
        "sheet_name": sheet_name,
        "header_row_index": None,
        "product_col_index": product_col + 1 if product_col is not None else None,
        "qty_col_index": None,
        "rows_parsed": len(parsed_rows),
        "rows": parsed_rows,
    }


@app.get("/")
def root():
    return {"status": "ok"}


@app.post("/import-xlsx")
async def import_xlsx(
    file: UploadFile = File(...),
    x_api_token: str | None = Header(default=None),
):
    check_token(x_api_token)

    content = await file.read()
    wb = load_workbook_safe(content)
    _sheet_name, rows = choose_best_sheet(wb)

    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")

    header_row_idx, header_map = detect_header_row(rows)

    if header_row_idx is not None and header_map:
        header_row = rows[header_row_idx]
        data_rows = rows[header_row_idx + 1 :]
        original_headers = [str(header_row[i] or "").strip() for i in range(len(header_row))]
    else:
        max_cols = max(len(r) for r in rows) if rows else 0
        data_rows = rows
        original_headers = [f"column_{i+1}" for i in range(max_cols)]

    normalized_headers = []
    used = set()

    for i, h in enumerate(original_headers):
        base = sql_identifier(h if h else f"column_{i+1}")
        col = base
        n = 2
        while col in used:
            col = f"{base}_{n}"
            n += 1
        used.add(col)
        normalized_headers.append(col)

    file_name = file.filename or "import.xlsx"
    table_name = f"import_{sql_identifier(os.path.splitext(file_name)[0])}"

    cleaned_rows = []
    row_counter = 0

    for row in data_rows:
        values = []
        has_non_empty = False

        for i in range(len(normalized_headers)):
            val = row[i] if i < len(row) else None
            if val is not None and str(val).strip() != "":
                has_non_empty = True
            values.append(None if val is None else str(val))

        if has_non_empty:
            row_counter += 1
            cleaned_rows.append((row_counter, values))

    libelle_source_col = find_best_libelle_column(normalized_headers)
    libelle_clean_created = False

    conn = get_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("create extension if not exists pg_trgm;")
                cur.execute("create extension if not exists unaccent;")

                cols_sql = ',\n'.join([f'"{c}" text' for c in normalized_headers])

                create_sql = f'''
                create table if not exists public."{table_name}" (
                  id bigserial primary key,
                  imported_at timestamptz default now(),
                  source_file text,
                  row_number bigint,
                  {cols_sql}
                );
                '''
                cur.execute(create_sql)

                insert_cols = ["source_file", "row_number"] + normalized_headers
                cols_list = ", ".join([f'"{c}"' for c in insert_cols])
                placeholders = ", ".join(["%s"] * len(insert_cols))

                insert_sql = f'''
                insert into public."{table_name}" ({cols_list})
                values ({placeholders})
                '''

                payload = []
                for row_number, row_values in cleaned_rows:
                    payload.append([file_name, row_number] + row_values)

                if payload:
                    cur.executemany(insert_sql, payload)

                if libelle_source_col:
                    cur.execute(f'''
                        alter table public."{table_name}"
                        add column if not exists libelle_clean text;
                    ''')

                    cur.execute(f'''
                        update public."{table_name}"
                        set libelle_clean = trim(
                            regexp_replace(
                                upper(unaccent("{libelle_source_col}")),
                                '[^A-Z0-9 ]+',
                                ' ',
                                'g'
                            )
                        )
                        where "{libelle_source_col}" is not null;
                    ''')

                    index_name = f'idx_{table_name}_libelle_clean_trgm'
                    cur.execute(f'''
                        create index if not exists "{index_name}"
                        on public."{table_name}"
                        using gin (libelle_clean gin_trgm_ops);
                    ''')

                    libelle_clean_created = True

    finally:
        conn.close()

    return JSONResponse({
        "success": True,
        "table_name": table_name,
        "source_file": file_name,
        "rows_imported": len(cleaned_rows),
        "columns": normalized_headers,
        "libelle_source_column_detected": libelle_source_col,
        "libelle_clean_created": libelle_clean_created,
    })


@app.post("/parse-client-xlsx")
async def parse_client_xlsx(
    file: UploadFile = File(...),
    x_api_token: str | None = Header(default=None),
):
    check_token(x_api_token)
    content = await file.read()
    result = parse_client_xlsx_bytes(content)
    return JSONResponse(result)


@app.post("/quote-jobs")
async def create_quote_job(
    x_api_token: str | None = Header(default=None),
    payload: dict = Body(...)
):
    check_token(x_api_token)

    supplier_id = payload.get("supplier_id")
    input_type = payload.get("input_type")
    payload_json = payload.get("payload_json", {})

    if input_type not in {"xlsx", "email_body"}:
        raise HTTPException(status_code=400, detail="input_type must be 'xlsx' or 'email_body'")

    conn = get_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    insert into public.quote_jobs (supplier_id, status, input_type, payload_json, attempt_count, max_attempts)
                    values (%s, 'queued', %s, %s, 0, 3)
                    returning id, status, created_at
                    """,
                    (supplier_id, input_type, json.dumps(payload_json))
                )
                row = cur.fetchone()

        return JSONResponse({
            "success": True,
            "job_id": row[0],
            "status": row[1],
            "created_at": row[2].isoformat() if row[2] else None,
        })
    finally:
        conn.close()


@app.get("/quote-jobs/{job_id}")
async def get_quote_job(
    job_id: int,
    x_api_token: str | None = Header(default=None),
):
    check_token(x_api_token)

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                select id, supplier_id, status, input_type, payload_json, attempt_count,
                       error_message, created_at, started_at, finished_at
                from public.quote_jobs
                where id = %s
                """,
                (job_id,)
            )
            row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Job not found")

        return JSONResponse({
            "id": row[0],
            "supplier_id": row[1],
            "status": row[2],
            "input_type": row[3],
            "payload_json": row[4],
            "attempt_count": row[5],
            "error_message": row[6],
            "created_at": row[7].isoformat() if row[7] else None,
            "started_at": row[8].isoformat() if row[8] else None,
            "finished_at": row[9].isoformat() if row[9] else None,
        })
    finally:
        conn.close()