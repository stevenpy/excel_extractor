import os
import time
import json
import base64
import re
import unicodedata
import zipfile
import smtplib
from io import BytesIO
from html import escape
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import psycopg
from psycopg import sql
from psycopg.rows import dict_row

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


DATABASE_URL = os.getenv("DATABASE_URL", "")
DEFAULT_SUPPLIER_CATALOG_TABLE = os.getenv("DEFAULT_SUPPLIER_CATALOG_TABLE", "")
SUPPLIER_CATALOG_TABLE_MAP = os.getenv("SUPPLIER_CATALOG_TABLE_MAP", "{}")

MIN_SIM = float(os.getenv("MIN_SIM", "0.4"))
STRONG_SIM = float(os.getenv("STRONG_SIM", "0.7"))

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USER)
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "true").lower() == "true"


def get_conn():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL not configured")
    return psycopg.connect(DATABASE_URL, sslmode="require")


def normalize_text(value):
    text = str(value or "").strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value):
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


PRODUCT_HEADER_CANDIDATES = [
    "type_de_fournitures_services",
    "type_fournitures_services",
    "fournitures_services",
    "fourniture",
    "produit",
    "product",
    "article",
    "designation",
    "libelle",
    "description",
    "item",
    "nom_produit",
    "prestation",
    "service",
]

QTY_HEADER_CANDIDATES = [
    "quantite",
    "quantity",
    "qty",
    "qte",
    "qte_",
    "nombre",
    "nb",
    "volume",
]


def load_workbook_safe(content: bytes):
    try:
        return load_workbook(filename=BytesIO(content), data_only=True)
    except (InvalidFileException, zipfile.BadZipFile):
        raise RuntimeError("Unsupported Excel format. The uploaded file is not a valid .xlsx file.")
    except Exception as e:
        raise RuntimeError(f"Unable to read Excel file: {e}")


def choose_best_sheet(wb):
    best_sheet_name = wb.sheetnames[0]
    best_rows = []
    best_score = -1

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        non_empty = 0
        text_cells = 0

        for row in rows[:300]:
            for cell in row:
                if cell is None or str(cell).strip() == "":
                    continue
                non_empty += 1
                if any(ch.isalpha() for ch in str(cell)):
                    text_cells += 1

        score = non_empty + text_cells * 2

        if score > best_score:
            best_score = score
            best_sheet_name = ws.title
            best_rows = rows

    return best_sheet_name, best_rows


def detect_header_row(rows):
    best_score = -1
    best_row_idx = None
    best_mapping = {}

    for row_idx, row in enumerate(rows[:40]):
        mapping = {}
        score = 0

        for col_idx, cell in enumerate(row):
            norm = normalize_key(cell)
            if not norm:
                continue

            mapping[col_idx] = norm

            if norm in PRODUCT_HEADER_CANDIDATES:
                score += 10
            if norm in QTY_HEADER_CANDIDATES:
                score += 5
            if len(norm) > 2:
                score += 1

        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_mapping = mapping

    if best_score < 5:
        return None, {}

    return best_row_idx, best_mapping


def find_column_by_candidates(header_map, candidates):
    for candidate in candidates:
        for col_idx, norm_name in header_map.items():
            if norm_name == candidate:
                return col_idx
    return None


def fallback_detect_product_col(data_rows):
    if not data_rows:
        return None

    max_cols = max(len(r) for r in data_rows)
    best_col = None
    best_score = -1

    for col_idx in range(max_cols):
        score = 0
        non_empty = 0
        long_text = 0

        for row in data_rows[:200]:
            val = row[col_idx] if col_idx < len(row) else None
            txt = normalize_text(val)
            if not txt:
                continue

            non_empty += 1
            if any(ch.isalpha() for ch in txt):
                score += 1
            if len(txt) >= 10:
                long_text += 1
                score += 2

        if non_empty >= 3 and long_text >= 3 and score > best_score:
            best_score = score
            best_col = col_idx

    return best_col


def is_footer_like_label(text: str) -> bool:
    txt = normalize_text(text)
    if not txt:
        return False

    txt_upper = txt.upper()

    footer_markers_exact = {
        "ETABLI A",
        "ETABLI A :",
        "LE",
        "LE :",
        "CACHET ET SIGNATURE DU CANDIDAT",
        "SIGNATURE DU CANDIDAT",
        "CACHET DU CANDIDAT",
    }

    footer_markers_contains = [
        "CACHET ET SIGNATURE",
        "SIGNATURE DU CANDIDAT",
        "ETABLI A",
        "REMISE GENERALE",
        "CONSENTIE SUR CATALOGUE",
    ]

    compact = re.sub(r"[^A-Z0-9: ]+", " ", txt_upper)
    compact = re.sub(r"\s+", " ", compact).strip()

    if compact in footer_markers_exact:
        return True

    if any(marker in compact for marker in footer_markers_contains):
        return True

    return False


def parse_client_xlsx_bytes(content: bytes):
    wb = load_workbook_safe(content)
    sheet_name, rows = choose_best_sheet(wb)

    if not rows:
        raise RuntimeError("Empty workbook")

    header_row_idx, header_map = detect_header_row(rows)
    parsed_rows = []

    if header_row_idx is not None:
        data_rows = rows[header_row_idx + 1:]
        product_col = find_column_by_candidates(header_map, PRODUCT_HEADER_CANDIDATES)
        qty_col = find_column_by_candidates(header_map, QTY_HEADER_CANDIDATES)

        if product_col is None:
            product_col = fallback_detect_product_col(data_rows)

        for idx, row in enumerate(data_rows, start=1):
            product_value = row[product_col] if product_col is not None and product_col < len(row) else None
            product_label = normalize_text(product_value)

            if not product_label:
                continue
            if is_footer_like_label(product_label):
                break

            low = product_label.lower()
            if "type de fournitures" in low:
                continue
            if "reference fournisseur" in low:
                continue
            if low == "pu":
                continue
            if len(product_label) < 3:
                continue

            quantity = 1
            if qty_col is not None and qty_col < len(row):
                raw_qty = row[qty_col]
                txt_qty = normalize_text(raw_qty).replace(",", ".")
                try:
                    q = float(txt_qty)
                    if q > 0:
                        quantity = q
                except Exception:
                    quantity = 1

            parsed_rows.append({
                "request_row_number": idx,
                "request_product_label": product_label,
                "product_label": product_label,
                "product_label_clean": normalize_text(product_label).upper(),
                "quantity": quantity,
            })

        return parsed_rows

    data_rows = rows
    product_col = fallback_detect_product_col(data_rows)

    for idx, row in enumerate(data_rows, start=1):
        product_value = row[product_col] if product_col is not None and product_col < len(row) else None
        product_label = normalize_text(product_value)

        if not product_label:
            continue
        if len(product_label) < 3:
            continue

        parsed_rows.append({
            "request_row_number": idx,
            "request_product_label": product_label,
            "product_label": product_label,
            "product_label_clean": normalize_text(product_label).upper(),
            "quantity": 1,
        })

    return parsed_rows


def parse_email_body(body: str):
    def clean_line(line):
        return re.sub(r"\s+", " ", str(line or "")).strip()

    def extract_quantity_and_label(line: str):
        txt = clean_line(line)

        # 48 x PRODUIT
        m = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*[xX*]\s+(.+?)\s*$", txt)
        if m:
            qty = float(m.group(1).replace(",", "."))
            label = clean_line(m.group(2))
            return qty, label

        # 48 - PRODUIT
        m = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*[-–—]\s+(.+?)\s*$", txt)
        if m:
            qty = float(m.group(1).replace(",", "."))
            label = clean_line(m.group(2))
            return qty, label

        # 48 PRODUIT
        m = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s+(.+?)\s*$", txt)
        if m:
            qty = float(m.group(1).replace(",", "."))
            label = clean_line(m.group(2))
            return qty, label

        return 1, txt

    lines = [clean_line(x) for x in str(body or "").splitlines()]
    lines = [x for x in lines if x]

    skip_exact = {
        "bonjour",
        "bonjour,",
        "bonsoir",
        "bonsoir,",
        "merci",
        "merci,",
        "cordialement",
        "cordialement,",
        "cdt",
        "cdlt",
        "salutations",
        "salutations,",
    }

    skip_contains = [
        "je voudrais",
        "je souhaite",
        "merci de me chiffrer",
        "merci de me faire parvenir",
        "demande de devis",
        "suite à l'appel",
    ]

    out = []
    idx = 0

    for line in lines:
        low = normalize_text(line).lower()

        if low in skip_exact:
            continue

        if any(expr in low for expr in skip_contains):
            continue

        if low.endswith(":"):
            continue

        if len(line) < 3:
            continue

        quantity, product_label = extract_quantity_and_label(line)

        product_label = normalize_text(product_label)
        if not product_label or len(product_label) < 3:
            continue

        idx += 1
        out.append({
            "request_row_number": idx,
            "request_product_label": product_label,
            "product_label": product_label,
            "product_label_clean": normalize_text(product_label).upper(),
            "quantity": quantity,
        })

    return out


def parse_supplier_table_map():
    try:
        value = json.loads(SUPPLIER_CATALOG_TABLE_MAP or "{}")
        if isinstance(value, dict):
            return value
    except Exception:
        pass
    return {}


def resolve_catalog_table(supplier_id: str | None):
    mapping = parse_supplier_table_map()

    if supplier_id and supplier_id in mapping:
        return mapping[supplier_id]

    if DEFAULT_SUPPLIER_CATALOG_TABLE:
        return DEFAULT_SUPPLIER_CATALOG_TABLE

    raise RuntimeError(
        "No catalog table configured. Set DEFAULT_SUPPLIER_CATALOG_TABLE "
        "or SUPPLIER_CATALOG_TABLE_MAP."
    )


def fetch_best_match(conn, table_name: str, request_row: dict):
    query_text = request_row.get("product_label_clean") or ""
    request_row_number = request_row.get("request_row_number")
    request_product_label = (
        request_row.get("request_product_label")
        or request_row.get("product_label")
        or ""
    )

    if not query_text:
        return None

    with conn.cursor(row_factory=dict_row) as cur:
        stmt = sql.SQL("""
            SELECT
              %s::int AS request_row_number,
              %s::text AS request_product_label,
              p.*,
              similarity(p.libelle_clean, %s) AS similarity_score
            FROM public.{table_name} p
            WHERE p.libelle_clean %% %s
            ORDER BY similarity_score DESC, length(p.libelle_clean) ASC
            LIMIT 1
        """).format(
            table_name=sql.Identifier(table_name)
        )

        cur.execute(
            stmt,
            (
                request_row_number,
                request_product_label,
                query_text,
                query_text,
            )
        )
        return cur.fetchone()


def pick_first(obj: dict, keys: list[str], fallback: str = ""):
    for key in keys:
        value = obj.get(key)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return fallback


def pick_number(obj: dict, keys: list[str], fallback: float = 0.0):
    for key in keys:
        value = obj.get(key)
        if value is not None and str(value).strip() != "":
            try:
                n = float(str(value).replace(",", ".").strip())
                if n == n:
                    return n
            except Exception:
                pass
    return fallback


def match_rows(conn, table_name: str, parsed_rows: list[dict]):
    merged_rows = []

    for req in parsed_rows:
        best = fetch_best_match(conn, table_name, req)

        merged = {**req}
        if best:
            merged.update(best)

        merged_rows.append(merged)

    return merged_rows


def build_email_result(items_in: list[dict]):
    total = 0.0
    found_count = 0
    not_found_count = 0

    found_rows = []
    not_found_rows = []

    for j in items_in:
        demande = pick_first(j, [
            "request_product_label",
            "product_label"
        ], "")

        reference = pick_first(j, [
            "code_produit",
            "reference_article",
            "reference",
            "ref_article",
            "code_article",
            "article",
            "gencod",
            "code_du_fournisseur"
        ], "")

        libelle_trouve = pick_first(j, [
            "libelle",
            "designation_article",
            "designation_produit",
            "designation",
            "description",
            "nom_du_catalogue_cofaq",
            "product_label"
        ], "")

        quantity = pick_number(j, ["quantity"], 1)
        tarif = pick_number(j, [
            "tarif_ht",
            "prix_ht",
            "prix_brut",
            "prix_net",
            "prix_vente_indicatif_ht"
        ], 0)

        sim = pick_number(j, ["similarity_score"], 0)

        has_catalog_match = reference != "" and libelle_trouve != ""
        is_accepted = has_catalog_match and sim >= MIN_SIM

        if is_accepted:
            line_total = quantity * tarif
            total += line_total
            found_count += 1

            statut = "À vérifier"
            if sim >= STRONG_SIM:
                statut = "Match fort"
            elif sim >= MIN_SIM:
                statut = "Match moyen"

            found_rows.append(f"""
                <tr>
                  <td style="border:1px solid #ddd;padding:10px;">{escape(demande)}</td>
                  <td style="border:1px solid #ddd;padding:10px;">{escape(reference)}</td>
                  <td style="border:1px solid #ddd;padding:10px;">{escape(libelle_trouve)}</td>
                  <td style="border:1px solid #ddd;padding:10px;text-align:center;">{int(quantity) if quantity == int(quantity) else quantity}</td>
                  <td style="border:1px solid #ddd;padding:10px;text-align:right;white-space:nowrap;">{tarif:.2f} €</td>
                  <td style="border:1px solid #ddd;padding:10px;text-align:right;white-space:nowrap;">{line_total:.2f} €</td>
                  <td style="border:1px solid #ddd;padding:10px;text-align:center;">{escape(statut)}</td>
                </tr>
            """)
        else:
            not_found_count += 1

            not_found_rows.append(f"""
                <tr>
                  <td style="border:1px solid #ddd;padding:10px;">{escape(demande)}</td>
                  <td style="border:1px solid #ddd;padding:10px;text-align:center;">{int(quantity) if quantity == int(quantity) else quantity}</td>
                  <td style="border:1px solid #ddd;padding:10px;text-align:center;">Non trouvé</td>
                </tr>
            """)

    total_row = ""
    if found_rows:
        total_row = f"""
            <tr style="background-color:#f5f5f5;font-weight:bold;">
              <td colspan="5" style="border:1px solid #ddd;padding:10px;text-align:right;">Total HT</td>
              <td style="border:1px solid #ddd;padding:10px;text-align:right;white-space:nowrap;">{total:.2f} €</td>
              <td style="border:1px solid #ddd;padding:10px;"></td>
            </tr>
        """

    found_rows_html = "".join(found_rows) + total_row
    not_found_rows_html = "".join(not_found_rows)

    not_found_section = ""
    if not_found_count > 0:
        not_found_section = f"""
          <h3 style="margin-top:28px;">Articles non retrouvés dans notre catalogue :</h3>
          <table style="border-collapse:collapse;width:100%;margin-top:10px;">
            <thead>
              <tr style="background-color:#f5f5f5;">
                <th style="border:1px solid #ddd;padding:10px;text-align:left;">Demande client</th>
                <th style="border:1px solid #ddd;padding:10px;text-align:center;">Quantité</th>
                <th style="border:1px solid #ddd;padding:10px;text-align:center;">Statut</th>
              </tr>
            </thead>
            <tbody>
              {not_found_rows_html}
            </tbody>
          </table>

          <p style="margin-top:12px;"><strong>Nombre de lignes non retrouvées :</strong> {not_found_count}</p>
        """

    html = f"""
    <div style="font-family:Arial,sans-serif;font-size:14px;color:#333;line-height:1.5;">
      <p>Bonjour,</p>
      <p>Merci pour votre demande de devis.</p>
      <p>Veuillez trouver ci-dessous notre retour sur les articles demandés :</p>

      <h3 style="margin-top:24px;">Articles retrouvés dans notre catalogue :</h3>
      <table style="border-collapse:collapse;width:100%;margin-top:10px;">
        <thead>
          <tr style="background-color:#f5f5f5;">
            <th style="border:1px solid #ddd;padding:10px;text-align:left;">Demande client</th>
            <th style="border:1px solid #ddd;padding:10px;text-align:left;">Référence</th>
            <th style="border:1px solid #ddd;padding:10px;text-align:left;">Libellé trouvé</th>
            <th style="border:1px solid #ddd;padding:10px;text-align:center;">Quantité</th>
            <th style="border:1px solid #ddd;padding:10px;text-align:right;">Prix unitaire HT</th>
            <th style="border:1px solid #ddd;padding:10px;text-align:right;">Total HT</th>
            <th style="border:1px solid #ddd;padding:10px;text-align:center;">Statut</th>
          </tr>
        </thead>
        <tbody>
          {found_rows_html}
        </tbody>
      </table>

      <p style="margin-top:12px;">
        <strong>Nombre de lignes retenues :</strong> {found_count}<br>
        <strong>Nombre de lignes non retrouvées :</strong> {not_found_count}<br>
        <strong>Total HT :</strong> {total:.2f} €
      </p>

      {not_found_section}

      <p style="margin-top:24px;">Cordialement,</p>
    </div>
    """

    return {
        "html": html,
        "found_count": found_count,
        "not_found_count": not_found_count,
        "total_ht": round(total, 2),
    }


def extract_recipient_email(payload: dict):
    email_from = payload.get("email_from")

    if isinstance(email_from, dict):
        value = email_from.get("value")
        if isinstance(value, list) and value:
            first = value[0]
            if isinstance(first, dict) and first.get("address"):
                return first.get("address")

        if email_from.get("address"):
            return email_from.get("address")

    if isinstance(email_from, str) and email_from.strip():
        return email_from.strip()

    return None


def build_reply_subject(payload: dict, found_count: int):
    original = str(payload.get("email_subject") or "demande de devis").strip()

    # retire un éventuel ancien préfixe RE / RE:
    original = re.sub(r"^\s*RE\s*:\s*", "", original, flags=re.IGNORECASE)
    original = re.sub(r"^\s*RE\s+", "", original, flags=re.IGNORECASE)

    return f"RE : {original} - {found_count} article(s) retenu(s)"


def send_email_html(to_email: str, subject: str, html_body: str):
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS or not SMTP_FROM:
        raise RuntimeError("SMTP configuration is incomplete")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM
    msg["To"] = to_email

    text_body = "Votre retour devis est disponible en version HTML."
    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        if SMTP_USE_TLS:
            server.starttls()
            server.ehlo()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_FROM, [to_email], msg.as_string())


def fetch_next_job():
    conn = get_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    with next_job as (
                      select id
                      from public.quote_jobs
                      where
                        (
                          status = 'queued'
                          or (
                            status = 'processing'
                            and visibility_deadline is not null
                            and visibility_deadline < now()
                          )
                        )
                        and attempt_count < max_attempts
                      order by created_at
                      for update skip locked
                      limit 1
                    )
                    update public.quote_jobs q
                    set status = 'processing',
                        started_at = coalesce(q.started_at, now()),
                        locked_at = now(),
                        visibility_deadline = now() + interval '10 minutes',
                        last_heartbeat_at = now(),
                        attempt_count = q.attempt_count + 1,
                        error_message = null
                    from next_job
                    where q.id = next_job.id
                    returning q.id, q.supplier_id, q.input_type, q.payload_json, q.attempt_count, q.max_attempts;
                """)
                row = cur.fetchone()
        return row
    finally:
        conn.close()


def mark_done(job_id: int, result_payload: dict):
    conn = get_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    update public.quote_jobs
                    set status = 'done',
                        finished_at = now(),
                        payload_json = %s,
                        locked_at = null,
                        visibility_deadline = null,
                        last_heartbeat_at = null
                    where id = %s
                """, (json.dumps(result_payload), job_id))
    finally:
        conn.close()


def mark_failed(job_id: int, error_message: str):
    conn = get_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    update public.quote_jobs
                    set
                        status = case
                            when attempt_count >= max_attempts then 'failed'
                            else 'queued'
                        end,
                        finished_at = case
                            when attempt_count >= max_attempts then now()
                            else null
                        end,
                        error_message = %s,
                        locked_at = null,
                        visibility_deadline = null,
                        last_heartbeat_at = null
                    where id = %s
                """, (error_message[:4000], job_id))
    finally:
        conn.close()

def heartbeat_job(job_id: int, extend_minutes: int = 10):
    conn = get_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    update public.quote_jobs
                    set last_heartbeat_at = now(),
                        visibility_deadline = now() + (%s || ' minutes')::interval
                    where id = %s
                      and status = 'processing'
                """, (extend_minutes, job_id))
    finally:
        conn.close()

def process_job(job_row):
    job_id, supplier_id, input_type, payload_json, attempt_count, max_attempts = job_row
    payload = payload_json or {}

    heartbeat_job(job_id)

    if input_type == "xlsx":
        file_b64 = payload.get("file_base64")
        if not file_b64:
            raise RuntimeError("Missing file_base64 in payload_json")
        content = base64.b64decode(file_b64)
        parsed_rows = parse_client_xlsx_bytes(content)
    elif input_type == "email_body":
        body = payload.get("email_body", "")
        parsed_rows = parse_email_body(body)
    else:
        raise RuntimeError(f"Unsupported input_type: {input_type}")

    heartbeat_job(job_id)

    recipient_email = extract_recipient_email(payload)
    if not recipient_email:
        raise RuntimeError("Unable to determine recipient email")

    catalog_table = resolve_catalog_table(supplier_id)

    conn = get_conn()
    try:
        matched_rows = match_rows(conn, catalog_table, parsed_rows)
    finally:
        conn.close()

    heartbeat_job(job_id)

    email_result = build_email_result(matched_rows)
    subject = build_reply_subject(payload, email_result["found_count"])

    send_email_html(
        to_email=recipient_email,
        subject=subject,
        html_body=email_result["html"]
    )

    result = {
        **payload,
        "supplier_id": supplier_id,
        "catalog_table": catalog_table,
        "rows_parsed": len(parsed_rows),
        "found_count": email_result["found_count"],
        "not_found_count": email_result["not_found_count"],
        "total_ht": email_result["total_ht"],
        "reply_subject": subject,
        "recipient_email": recipient_email,
    }

    mark_done(job_id, result)


def main():
    while True:
        job = fetch_next_job()

        if not job:
            time.sleep(2)
            continue

        job_id = job[0]

        try:
            process_job(job)
            print(f"Job {job_id} done")
        except Exception as e:
            print(f"Job {job_id} failed: {e}")
            mark_failed(job_id, str(e))


if __name__ == "__main__":
    main()